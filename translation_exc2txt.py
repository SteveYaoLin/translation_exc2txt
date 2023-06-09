import openpyxl
import os

def extract_row_elements(filepath, sheetname, row_index, output_filename):
    # 加载Excel文件
    workbook = openpyxl.load_workbook(filepath)
    # 选择工作表
    sheet = workbook[sheetname]
    # 获取指定行的所有单元格
    row = sheet[row_index]
    # 获取Python脚本所在的目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # 构建输出文件路径
    output_filepath = os.path.join(script_dir, output_filename)
    # 检查是否存在同名文件，如果存在，则递增数字命名
    count = 1
    while os.path.exists(output_filepath):
        output_filename = f"output{count}.txt"
        output_filepath = os.path.join(script_dir, output_filename)
        count += 1
    # 创建文本文件
    with open(output_filepath, 'w') as output_file:
        # 逐个写入每个单元格的值，每个值占据一行
        for cell in row:
            output_file.write(str(cell.value) + '\n')

    print(f"提取完成！输出文件为: {output_filepath}")

# 获取用户输入的Excel文件路径
excel_filepath = input("请输入Excel文件路径：")
# 加载Excel文件
workbook = openpyxl.load_workbook(excel_filepath)
# 获取所有工作表的名称
sheet_names = workbook.sheetnames

print("备选的工作表：")
for i, sheet_name in enumerate(sheet_names):
    print(f"{i+1}. {sheet_name}")

# 获取用户选择的工作表索引
selected_sheet_index = int(input("请输入要处理的工作表的索引：")) - 1
# 根据索引获取工作表名称
selected_sheet_name = sheet_names[selected_sheet_index]

# 获取用户选择的行数
selected_row_index = int(input("请输入要转换的行数："))

# 指定输出文件名
output_filename = 'output.txt'

extract_row_elements(excel_filepath, selected_sheet_name, selected_row_index, output_filename)
